#! /usr/bin/python
#
# Estimate Scope_Items Export
#
__author__ = u'morrj140'
__VERSION__ = u'0.1'

import datetime
import time

from Logger import *
logger = setupLogging(__name__)
logger.setLevel(INFO)

from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl.worksheet import Worksheet as worksheet

from ArchiLib import ArchiLib
from Neo4JLib import Neo4JLib
from Constants import *

import pytest

class CreateEstimate(object):
    fileExcelIn  = None
    fileExcelOut = None

    def __init__(self, gdb):
        self.fileExcelIn = fileExcelIn
        self.fileExcelOut = fileExcelOut

        self.nj = Neo4JLib(gdb)

    def query(self):
        qs = u"MATCH "
        qs = qs +    u"(n0:ApplicationFunction)-- (r0)"
        qs = qs + u"-- (n1:ApplicationComponent)--(r1)"
        qs = qs + u"-- (n2:ApplicationService)--  (r2)"
        qs = qs + u"-- (n3:BusinessProcess)--     (r3)"
        qs = qs + u"-- (n4:BusinessObject) "
        qs = qs + u"Return n0, r0, n1, r1, n2, r2, n3, r3, n4, n4.PageRank, n4.RequirementCount, n4.Degree"

        self.lq, self.qd = self.nj.cypherQuery(qs)

        self.queryExportExcel(self.lq)

        logger.info(u"%d rows returned" % len(self.lq))

    def exportExcel(self):

        self.query()

        wb = load_workbook(filename = self.fileExcelIn)

        wsTitle = u"Import Items" + time.strftime(u"%Y%d%m_%H%M%S")

        ws = wb.create_sheet()
        ws.title = wsTitle

        logger.info(u"Created Worksheet %s" % wsTitle)

        n = 0
        for x in self.lq:
            n += 1

            logger.debug(u"x in lq : %s" % x)

            m = 0
            for y in x:
                m += 1

                col = get_column_letter(m)
                logger.debug(u"col : %s" % col)

                logger.debug(u"y : %s" % y)

                rs = u"%s%s" % (col, n)
                logger.debug(u"%s : %s" % (rs, y))

                ws.cell(rs).value = (u"%s" % (y))

        wb.save(filename=fileExcelOut)

        logger.info(u"Saved file : %s" % fileExcelOut)

def CreateEstimate():
    start_time = ArchiLib.startTimer()

    ce = CreateEstimate(gdbTest)

    ce.exportExcel()

    ArchiLib.stopTimer(start_time)

if __name__ == u"__main__":
    CreateEstimate()