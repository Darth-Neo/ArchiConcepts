#!/usr/bin/python
#
# Neo4J Counts
#
__author__  = u'morrj140'
__VERSION__ = u'0.1'

from Constants import *
from Logger import *
logger = setupLogging(__name__)
logger.setLevel(INFO)

from nl_lib.Concepts import Concepts
from py2neo import neo4j, node, rel
from py2neo.neo4j import Node
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl.worksheet import Worksheet as worksheet

import pytest


class Neo4JLib(object):

    _delchars = ''.join(c for c in map(chr, range(255)) if (not c.isalnum() and c != ' '))

    def __init__(self, gdb, fileCSVExport=fileCSVExport):
        logger.debug(u"Neo4J instance : %s" % gdb)

        self.graph = neo4j.GraphDatabaseService(gdb)

        self.fileCSVExport = fileCSVExport

    def _cleanString(self, name):
            '''
            :param name: name
            :return: string
            Note: Encode a string to UTF-8
            Note: Decode a string to Unicode
            '''
            name = name.replace(u",", u"-")
            name = name.replace(u")", u" ")
            name = name.replace(u"(", u"-")

            try:
                if isinstance(name, (str, unicode)):
                    return unicode(name)
                else:
                    n = unicode(name).strip()
                    return n
            except:
                return u" "

    def cypherQuery(self, qs):
        query = neo4j.CypherQuery(self.graph, qs)
        return query.execute().data


    def traversal(self, ql, directed=False):

        qs = u"MATCH"
        n=0

        if directed is True:
            sRelation = u"--[r%d]->"
        else:
            sRelation = u"-[r%d]-"

        lenRelation = len(sRelation % n)

        for x in ql:
            sAddRelation = sRelation % n
            qs = u"%s(n%d:%s)%s" % (qs, n, x, sAddRelation)
            n += 1

        qr = u" Return"
        for m in range(0, n, 1):
            qr = u"%s n%s, r%d," % (qr, m, m)

        query = qs[:-lenRelation] + qr[:-5]

        logger.info(u"%s" % query)

        return query

    def cypherQuery(self, qs):

        query = neo4j.CypherQuery(self.graph, qs)

        qd = query.execute().data

        listQuery = list()

        for x in qd:

            xl = list()

            for v in x.values:

                logger.debug(u"Type : %s" % type(v))

                if isinstance(v, Node):
                    name = v[u"aname"]
                    typeName =v[u"typeName"]

                    xl.append(name)
                    # xl.append(typeName)

                elif isinstance(v, int) or isinstance(v, float):
                    count = v
                    typeName = u"Number"

                    xl.append(count)
                    #xl.append(typeName)

                elif isinstance(v, str) or isinstance(v, unicode):
                    st = v
                    typeName = u"String"

                    xl.append(st)
                    # xl.append(typeName)

            listQuery.append(xl)

        return listQuery, qd

    def _cypherQuery(self, qs):

        query = neo4j.CypherQuery(self.graph, qs)

        qd = query.execute().data

        listQuery = list()

        n = 0

        ColHdr = True

        for x in qd:

            if n == 0:
                n += 1

                logger.debug(u"%s[%d]" % (x, n))

                cl = list()
                m = 0
                for y in x.columns:
                    m += 1
                    logger.debug(u"  %s[%d]" % (y, m))
                    cl.append(y)
                    cl.append(u"Type%d" % m)

                listQuery.append(cl)

            xl = list()

            for v in x.values:

                logger.debug(u"Type : %s" % type(v))

                if isinstance(v, Node):
                    name = v[u"name"]
                    typeName = v[u"typeName"]

                    xl.append(name)
                    xl.append(typeName)

                elif isinstance(v, int) or isinstance(v, float):
                    count = v
                    typeName = u"Number"

                    xl.append(count)
                    xl.append(typeName)

                elif isinstance(v, str) or isinstance(v, unicode):
                    st = v
                    typeName = u"String"

                    xl.append(st)
                    xl.append(typeName)

            listQuery.append(xl)

        return listQuery, qd

    def logResults(self, lq, f=None, n=0):
        n += 1

        spaces = u" " * n

        rs = u""
        for x in lq:
            if isinstance(x, tuple) or isinstance(x, list):
                self.logResults(x, f, n)

            elif isinstance(x, str) or isinstance(x, unicode):
                if x == u"Nodes" or x is None or len(x) == 0:
                    continue
                logger.debug(u"%s %s" % (spaces, x))
                x = self.cleanString(x)
                rs += u", \"%s\"" % (x)

            elif isinstance(x, float) or isinstance(x, int):
                logger.debug(u"%s %s" % (spaces, x))
                rs += u", %s" % (x)

        if len(rs) != 0:
            rs1 = rs[1:]

            u'''
            :param name: name
            :return: string
            Note: Encode a string to UTF-8
            Note: Decode a string to Unicode
            '''
            # rsClean = self.cleanString(rs1)
            # unicode(n, "utf-8", errors='ignore' )

            logger.debug(u"%s" % (rs1))

            f.write(u"%s\n" % (rs1))

    def queryExport(self, lq):

        # csvExport defined in al_ArchiLib.Constants
        f = open(self.fileCSVExport, u'w')

        self.logResults(lq, f)

        f.close()

        logger.info(u"Exported %d rows" % len(lq))
        logger.info(u"Save Model : %s" % self.fileCSVExport)

    def neo4jCounts(self):

        qs = u"MATCH (n) RETURN n.typeName, count(n.typeName) as Count order by Count DESC"
        lq, qd = self.cypherQuery(qs)

        logger.info(u"Neo4J Counts")

        sl = sorted(lq[1:], key=lambda c: int(c[1]), reverse=True)
        for x in sl:
            if len(x) == 2:
                logger.debug(u"%4d : %s" % (x[1], x[0]))
            else:
                logger.error(u"List of wrong length : %d" % len(x))

        return sl

    def queryExportExcel(self, lq, fileIn=fileExcelIn, fileOut=fileExcelOut):

        wb = load_workbook(filename=fileIn)

        wsTitle = u"Import Items" + time.strftime(u"%Y%d%m_%H%M%S")

        ws = wb.create_sheet()
        ws.title = wsTitle

        logger.info(u"Created Worksheet %s" % wsTitle)

        n = 0
        for x in lq:
            n += 1

            logger.debug(u"x in lq : %s" % x)

            m = 0
            for y in x:
                m += 1

                col = get_column_letter(m)
                logger.debug(u"col : %s" % col)

                logger.debug(u"y : %s" % y)

                rs = u"%s%s" % (col, n)
                logger.debug(u"%s : %s" % (rs, y))

                ws.cell(rs).value = (u"%s" % (y))

        try:
            wb.save(fileOut)

        except:
            pass

        logger.info(u"Saved file : %s" % fileOut)

    def exportNeo4JToConcepts(self, concepts, fileNodes=u"nodes.p"):
        # #et all nodes
        # Match n return n limit 25

        qs = u"Match n return n"

        lq, qd = self.cypherQuery(qs)

        for x in lq:
            if len(x) == 2:
                logger.info(u"%s[%s]" % (x[0], x[1]))
                concepts.addConceptKeyType(x[0], x[1])
            else:
                logger.warn(u"Not a standard node : %s" % x)

        # Match r relations
        qs = u"match n-[r]-m return n, r, m"
        lq, qd = self.cypherQuery(qs)

        for x in lq:
            if len(x) == 6:
                logger.info(u"%s[%s]" % (x[0], x[1]))
                concepts.addConceptKeyType(x[0], x[1])
            else:
                logger.warn(u"Not a standard node : %s" % x)

        Concepts.saveConcepts(concepts, fileNodes)

        return concepts

@pytest.fixture(scope="module")
def gdb():
    return gdbTest

@pytest.fixture(scope="module")
def nj(gdb):
    return Neo4JLib(gdb)

@pytest.mark.NeoJ
def test_Neo4jLib(nj):

    sl = nj.neo4jCounts()

    assert(sl is not None)


if __name__ == u"__main__":
    test_Neo4jLib()