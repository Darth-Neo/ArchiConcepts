#!/usr/bin/python
#
# Archimate Libray
#
__author__ = u'morrj140'
__VERSION__ = u'0.1'

import sys
import os
import StringIO
import csv
import random
import time

from Logger import *

logger = setupLogging(__name__)
logger.setLevel(INFO)

from nl_lib.Concepts import Concepts

from lxml import etree

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter

from Constants import *


#
# Main class to make life easier
#
class ArchiLib(object):
    dictName  = dict()
    dictEdges = dict()
    dictNodes = dict()
    dictBP    = dict()
    dictCount = dict()

    def __init__(self, fileArchimate):

        if fileArchimate != None:
            self.fileArchimate = fileArchimate
        else:
            self.fileArchimate = fileArchimate

        etree.QName(ARCHIMATE_NS, u'model')

        self.tree = etree.parse(self.fileArchimate)

        # Populate Dictionaries for easier code
        self.parseAll()

    def outputXMLtoFile(self, filename=u"import_artifacts.archimate"):
        output = StringIO.StringIO()
        self.tree.write(output, pretty_print=True)

        logger.debug(u"%s" % (output.getvalue()))

        logger.info(u"====> Saved to : %s" % filename)

        with open(filename, u'w') as f:
            f.write(output.getvalue())

        output.close()

    def outputCSVtoFile(self, concepts, fileExport):

        self.fileExport = fileExport

        listOutput = concepts.listCSVConcepts()

        colDict = dict()

        f = open(self.fileExport, u'w')

        m = 0
        for x in listOutput:
            m += 1
            n = 0
            strLine = u""
            logger.debug(u"listOutput[%d] = %s" % (n, x))

            for y in x.split(u","):
                n += 1

                logger.debug(u"y : %s[%d]" % (y, len(y)))

                if len(y) == 0:
                    if colDict.has_key(n):
                        y = colDict[n]
                else:
                    colDict[n] = y

                strLine = u"%s%s," % (strLine, y)

            nl = strLine[:-1]

            logger.debug(u"%s" % nl)
            f.write(nl + u"\n")

        f.close()
        logger.info(u"Save Model : %s" % self.fileExport)

    def outputXMLtoLog(self):
        output = StringIO.StringIO()
        self.tree.write(output, pretty_print=True)
        logger.info(u"%s" % (output.getvalue()))

    def exportExcel(fileIn, fileOut, workSheetTitle=u"Scope Items"):

        wb = load_workbook(filename = fileIn)

        ws = wb.create_sheet()

        ws.title = workSheetTitle

        ws[u'F5'] = 3.14

        for col_idx in range(1, 40):
            col = get_column_letter(col_idx)
            for row in range(1, 600):
                ws.cell(u'%s%s'%(col, row)).value = u'%s%s' % (col, row)

        wb.save(filename = fileOut)


    #
    # Model transversal functions via XPath
    #
    def findDiagramModelByName(self, name):
        r = None

        xp = u"//element[@name='" + name + u"']"
        stp = self.tree.xpath(xp)

        if stp[0].get(ARCHI_TYPE) == DIAGRAM_MODEL:
            r = stp[0]

        return r

    def findRelationsByID(self, id, Target=False):
        if Target is False:
            xp = u"//element[@source='%s']" % (id)
        else:
            xp = u"//element[@source='%s' or @target='%s']" % (id, id)
        logger.debug(u"%s" % xp)
        stp = self.tree.xpath(xp)
        return stp

    def findRelationsByTargetID(self, id):
        xp = u"//element[@target='%s']" % id
        stp = self.tree.xpath(xp)
        return stp

    def findDiagramModel(self, id):
        xp = u"//element[@id='" + id + "']"
        stp = self.tree.xpath(xp)
        return stp

    def findDiagramModelByName(self, name):
        xp = u"//element[@name='%s']" % (name)
        logger.debug(u"xp : %s" % xp)
        stp = self.tree.xpath(xp)

        for x in stp:
            if x.get(ARCHI_TYPE) == DIAGRAM_MODEL:
                stp = x
                break

        return stp

    def findDiagramObject(self, id):
        xp = u"//child[@id='%s']" % id
        stp = self.tree.xpath(xp)
        return stp

    def findElement(self, name):
        xp = u"//element[@name='%s']" % name.rstrip(u" ").lstrip(u" ")
        stp = self.tree.xpath(xp)
        return stp

    def findElementByID(self, id):
        xp = u"//element[@id='%s']" % id
        stp = self.tree.xpath(xp)

        if stp is None:
            stp = list()
            stp.append(u" ")

        return stp

    def getDiagramModels(self):
        xp = u"//element[@%s='%s']" % (ARCHI_TYPE, DIAGRAM_MODEL)
        stp = self.tree.xpath(xp)
        return stp

    def recurseElement(self, e, concepts, n=0):

        n += 1

        try:
            attributes = e.attrib
        except:
            logger.warn(u"Ops...")
            return concepts

        logger.debug(u"recurseElement %s: %s:%s:%s:%s" % (n, e.tag, e.get(u"name"), e.get(u"id"), attributes.get(ARCHI_TYPE)))

        if attributes.get(u"id") is not None:
            id = attributes[u"id"]
            name = e.get(u"name")
            type = e.get(ARCHI_TYPE)[10:]

            d = concepts.addConceptKeyType(name, type)

            next = self.findRelationsByID(id)

            for x in next:
                nid = x.get(u"target")
                en = self.findElementByID(id)
                self.recurseElement(en, d, n)

        return concepts


    def recurseChildren(self, concepts, x, n=0):
        n += 1

        xid = x.get(u"id")
        xi = x.items()
        xc, xname = self.getElementName(x.get(u"archimateElement"))

        ce = self.tree.xpath(u"//child[@id='%s']" % (xid))
        nc  = ce[0].getchildren()

        #
        # Add source into concepts
        #
        #c = concepts.addConceptKeyType(ModelToExport + ":" + xname, x.get(ARCHI_TYPE)[10:])

        #
        # for each RelationShip, find the Source and Target
        #
        for y in nc:
            if y.tag == u"child":
                logger.debug(u"%d.%s" % (n, y.tag))
                yc, yname = self.getElementName(y.get(u"archimateElement"))
                d = concepts.addConceptKeyType(yname, yc.get(ARCHI_TYPE)[10:])
                self.recurseChildren(d, y)

            if y.tag == u"sourceConnection":
                logger.debug(u"skip - %s" % y.tag)

                yid = y.get(u"id")
                yi = y.items()

                sid = y.get(u"source")
                s, sname = self.getChildName(sid)

                tid = y.get(u"target")
                t, tname = self.getChildName(tid)

                relid = y.get(u"relationship")
                rel, rname = self.getElementName(relid)

                if rname is None:
                    rname = u"Target"

                d = concepts.addConceptKeyType(rname, rel.get(ARCHI_TYPE)[10:])
                e = d.addConceptKeyType(tname, t.get(ARCHI_TYPE)[10:])

                logger.debug(u"    %s" % yi)
                logger.debug(u"%s,%s,%s,%s,%s,%s\n" % (sname, s.get(ARCHI_TYPE), rel.get(ARCHI_TYPE), rname, tname, t.get(ARCHI_TYPE)))

            # recurseElement(t, e, tree)

    def recurseModel(self, ModelToExport, concepts):
        #
        # Find DiagramModel Element to Export
        #
        xp = u"//element[@name='%s']" % (ModelToExport)
        logger.debug(u"XP : %s" % xp)

        se = self.tree.xpath(xp)

        nse = len(se)
        logger.debug(u"num se : %d" % nse)
        #
        # Iterate over the DiagramModel's DiagramObjects children
        #
        for m in se:
            if m.get(ARCHI_TYPE) == DIAGRAM_MODEL:
                logger.debug(u"%s:%s:%s" % (m.get(u"name"), m.get(ARCHI_TYPE), m.tag))

                r = m.getchildren()

                c = concepts.addConceptKeyType(m.get(u"name"), m.get(ARCHI_TYPE)[10:])

                #
                # for each DiagramObject, Find the ArchimateElement
                #
                for x in r:

                    if x.get(ARCHI_TYPE) is not u"archimate:DiagramObject":
                        continue

                    xid = x.get(u"id")

                    xc, xname = self.getElementName(x.get(u"archimateElement"))

                    logger.info(u"  %s[%s]" % (xname, x.get(ARCHI_TYPE)))

                    d = c.addConceptKeyType(xname, xc.get(ARCHI_TYPE)[10:])

                    self.recurseChildren(d, x)

        return concepts

    #
    # Model functions via dictionaries
    #
    def logTypeCounts(self, ListOnly=False):
        if not ListOnly:
            logger.info(u"Type Counts")

        listCounts = self.dictCount.items()

        if not ListOnly:
            for x in sorted(listCounts, key=lambda c: abs(c[1]), reverse=True):
                if x[1] > 1:
                    logger.info(u"  %d - %s" % (x[1], x[0]))

            logger.info(u" ")

        return listCounts

    def countNodeType(self, type):
        if self.dictCount.has_key(type):
            self.dictCount[type] += 1
        else:
            self.dictCount[type] = 1

    def parseAll(self):
        for x in self.tree.getroot():
            logger.debug(u"Folder : %s" % x.get(u"name"))
            self.getElementsFromFolder(x.get(u"name"))

    def getElementsFromFolder(self, folder):

        se = self.tree.xpath(u"folder[@name='%s']" % (folder))

        if folder == u"Views":
            return se[0]

        if folder == u"Relations":
            for x in se:
                self.getNode(x, self.dictEdges)
        else:
            for x in se:
                self.getNode(x, self.dictNodes)

        return se

    def getNode(self, el, d):
        logger.debug(u"el.tag = %s" % (el.tag))

        attributes = el.attrib

         # Not every node will have a type
        if el.tag in (u"element", u"child"):
            self.countNodeType(attributes[ARCHI_TYPE])

        nl = dict()
        for atr in attributes:
            nl[atr] = attributes[atr]
            logger.debug(u"%s = %s" % (atr, attributes[atr]))

        if nl.has_key(u"id"):
            d[nl[u"id"]] = nl
            self.dictName[el.get(u"name")] = el.get(u"id")

        for elm in el:
            self.getNode(elm, d)

    def getTypeNodes(self, type):
        d = dict()

        for x in self.dictNodes.values():
            xt = x.get(ARCHI_TYPE)
            if xt in type:
                d[x.get(u"name")] = x
        return d

    def getNodeName(self, node):
        name = u" "

        try:
            logger.debug(u"  Node : %s" % (self.dictNodes[node][u"name"]))
            name = self.dictNodes[node][u"name"]
        except:
            logger.debug(u"Node not Found")

        return name

    def getEdgesForNode(self, source, searchType, n=5, m=0):
        listNodes = list()

        if n == 0:
            return listNodes
        else:
            n -= 1

        m += 1

        xp = u"//element[@source='%s']" % source
        stp = self.tree.xpath(xp)

        for x in stp:
            logger.debug(u"Index %d:%d Depth" % (m, n))

            sourceNE = x.get(u"source")
            targetNE = x.get(u"target")

            dns = x.get(ARCHI_TYPE)

            if  dns in searchType:
                spaces = u" " * n
                nodeName = self.getNodeName(targetNE)
                if nodeName != u"NA":
                    nn = u"%s%s" % (spaces, nodeName)
                    listNodes.append(nn)

                    ln = self.getEdgesForNode(targetNE, searchType, n, m)
                    for y in ln:
                        listNodes.append(y)

        return listNodes


    def getModelsInFolder(self, folder):
        xp = u"//folder[@name='%s']" % (folder)

        se = self.tree.xpath(xp)

        modelList = se[0].getchildren()

        models = list()

        for x in modelList:
            modelName = str(x.get(u"name"))
            models.append(modelName)

        return models

    def getFolders(self):
        se = self.tree.xpath(u'//folder')

        l = list()

        for x in se:
            l.append(x.get(u"name"))
            logger.debug(u"%s" % (x.get(u"name")))

        return l

    def getChildName(self, id):

        xp = u"//child[@id='%s']" % str(id)

        logger.debug(u"xp : %s" % xp)

        se = self.tree.xpath(xp)

        if len(se) > 0:
            ae = se[0].get(u"archimateElement")
            return self.getElementName(ae)

    def getElementName(self, id):

        xp = u"//element[@id='%s']" % str(id)

        logger.debug(u"xp : %s" % xp)

        se = self.tree.xpath(xp)

        if len(se) > 0:
            element = se[0]
            name = se[0].get(u"name")
            if name is None:
                name = element.get(ARCHI_TYPE)[10:]

            logger.debug(u"getElementName - %s:%s" % (element.get(ARCHI_TYPE), name))
            return element, name
        else:
            return None, None

    def getNameID(self, value):
        logger.info(u"    Search for : %s" % value)
        if self.dictName.has_key(value):
            idd = self.dictName[value]
            logger.debug(u"    Found! : %s" % idd)
        else:
            idd = self._getID()
            self.dictName[value] = idd
            logger.debug(u"New I  : %s" % idd)

        logger.debug(u"%s" % self.dictName)

        return idd

    #
    # Model functions via dictionaries
    #
    def logTypeCounts(self):
        listCounts = list()

        logger.info(u"Type Counts")
        listCounts = self.dictCount.items()

        for x in sorted(listCounts, key=lambda c: abs(c[1]), reverse=True):
            if x[1] > 1:
                logger.info(u" %d - %s" % (x[1], x[0]))

        logger.info(u" ")

        return listCounts

    #
    # Node - <element xsi:type="archimate:Node" id="612a9b73" name="Linux Server"/>
    #
    def insertNode(self, tag, folder, attrib):
        idd = None

        try:
            logger.debug(u"attrib: %s" % (attrib))

            value = attrib[u"name"].rstrip(u" ").lstrip(u" ")

            if value != attrib[u"name"]:
                logger.debug(u"diff value .%s:%s." % (value, attrib[u"name"]))

            if self.dictName.has_key(value):
                idd = self.dictName[value]
                attrib[u"id"] = idd

                logger.debug(u"inFound! : %s" % idd)
            else:
                idd = self._getID()
                self.dictName[value] = idd
                attrib[u"id"] = idd

                elm = etree.Element(tag, attrib, nsmap=NS_MAP)

                xp = u"//folder[@name='" + folder + u"']"
                txp = self.tree.xpath(xp)
                txp[0].insert(0, elm)
                logger.debug(u"uinNew!   : %s" % idd)

        except:
            logger.warn(u"attrib: %s" % (attrib))

        return idd

    def insertRel(self, tag, folder, attrib):

        logger.debug(u"attrib: %s" % (attrib))

        value = u"%s--%s" % (attrib[u"source"], attrib[u"target"])

        if self.dictName.has_key(value):
            idd = self.dictName[value]
            attrib[u"id"] = idd

            logger.debug(u"inFound! : %s" % idd)
        else:
            idd = self._getID()
            self.dictName[value] = idd
            attrib[u"id"] = idd

            xp = u"//folder[@name='" + folder + u"']"
            elm = etree.Element(tag, attrib, nsmap=NS_MAP)
            self.tree.xpath(xp)[0].insert(0, elm)
            logger.debug(u"inNew!   : %s" % idd)

        return idd

    # Properties
    # <element xsi:type="archimate:BusinessProcess" id="0ad0bac9" name="06.0 Activity Reports">
    #        <property key="ExampleName" value="ExampleValue"/>
    # </element>
    def addProperties(self, properties):

        # To Do - stop duplicating properties

        idd = properties[u"ID"]
        node = self.findElementByID(idd)

        n = 0
        for key, value in properties.items():
            if key != u"ID"and node.get(key) is None:
                prop = dict()
                prop[u"key"] = key
                prop[u"value"] = value
                elm = etree.Element(u"property", prop, nsmap=NS_MAP)
                node[0].insert(n, elm)
                n += 1

    def insertNColumns(self, folder, subfolder, fileMetaEntity):

        file = open(fileMetaEntity, u"rU")
        reader = csv.reader(file)

        xp = u"folder[@name='" + folder + u"']"
        tag = u"element"

        # <folder name="Process" id="e23b1e50">

        attrib = dict()
        attrib[u"id"] = self._getID()
        attrib[u"name"] = subfolder
        self.insertNode(u"folder", folder, attrib)

        folder = subfolder

        rownum = 0

        previous = dict()

        listColumnHeaders = list()

        properties = dict()

        PROPERTIES_FLAG = False

        for row in reader:

            if rownum == 0:
                rownum += 1
                for col in row:
                    if col[:8] == u"Property":
                        colType = col
                        listColumnHeaders.append(colType)
                    else:
                        colType = u"archimate:%s" % col
                        listColumnHeaders.append(colType)
                continue

            logger.info(u"----------------------------------------------------------------------------------------")
            logger.debugu(u"rownum : %d" % rownum)
            logger.debug(u"row    : %s" % row)

            p = None
            colnum = 0

            for col in row:
                logger.info(u"    %d   [%s] %s" % (colnum, listColumnHeaders[colnum], col))

                CM = unicode(col).lstrip()

                if listColumnHeaders[colnum][:8] == u"Property":
                    logger.debug(u"Properties : %s - %s" % (listColumnHeaders[colnum][9:], CM))

                    if u"ID" not in properties:
                        properties[u"ID"] = p

                    properties[listColumnHeaders[colnum][9:]] = CM

                    colnum += 1
                    continue

                if len(properties) > 0:
                    logger.debug(u"Add %d Properties" % (len(properties)))
                    self.addProperties(properties)
                    properties = dict()

                #
                # This is for a cvs which assumes value in column
                # from a previous column
                #
                if CM == u"" or CM is None:
                    logger.debug(u"Using %d[%s]" % (colnum, previous[colnum]))
                    CM = previous[colnum]
                else:
                    previous[colnum] = CM
                    logger.debug(u"CM  %d[%s]" % (colnum, CM))

                #
                # Create the attributes
                #
                attrib = dict()
                attrib[u"name"] = CM
                attrib[ARCHI_TYPE] = listColumnHeaders[colnum]
                self.insertNode(tag, folder, attrib)
                CM_ID = attrib[u"id"]

                if p is not None:
                    attrib = dict()
                    attrib[u"source"] = CM_ID
                    attrib[u"target"] = p
                    attrib[ARCHI_TYPE] = u"archimate:AssociationRelationship"
                    self.insertRel(tag, u"Relations", attrib)

                    p = CM_ID
                else:
                    p = CM_ID

                colnum += 1

        if len(properties) > 0:
            logger.info(u"Add %d Properties" % (len(properties)))
            self.addProperties(properties)
            properties = dict()

    def insertConcepts(self, tree, concepts, n=0):

        for x in concepts.getConcepts().values():
            logger.info(u"x : %s" % x.name)
            for y in x.getConcepts().values():
                logger.info(u"  y : %s" % y.name)
                attrib = dict()
                attrib[u"name"] = x.name
                attrib[ARCHI_TYPE] = u"archimate:WorkPackage"
                self.insertNode(u"element", u"Implementation & Migration", attrib)
                wp1 = attrib[u"id"]

                attrib = dict()
                attrib[u"name"] = y.name
                attrib[ARCHI_TYPE] = u"archimate:BusinessProcess"
                self.insertNode(u"element", u"Process", attrib)
                wp2 = attrib[u"id"]

                attrib = dict()
                attrib[u"source"] = wp1
                attrib[u"target"] = wp2
                attrib[ARCHI_TYPE] = u"archimate:AssociationRelationship"
                self.insertRel(u"element", u"Relations", attrib)

    #
    # These functions interact with nl_lib
    #

    def folderConcepts(self, concepts):
        r = self.tree.xpath(u'folder')

        for x in r:

            folder = str(x.get(u"name")).strip()

            logger.debug(u"folder : %s" % (folder))

            se = self.tree.xpath(u"folder[@name='%s']" % (folder))

            for element in se:
                self.createConcepts(concepts, element)

        # concepts.logConcepts()


    def conceptAttributes(self, c, el, n):
        n += 1
        spaces = " " * n

        attrib = el.attrib

        d = c.addConceptKeyType(u"Attributes", u"Attribute")

        attributes = el.attrib
        for atr in attributes.keys():
            logger.info(u"%sAttributes[%s]=%s" % (spaces, atr, attributes[atr]))
            d.addConceptKeyType(atr, attributes[atr])

        if el.tag == u'Documentation':
            d.addConceptKeyType(el.text, u"Text")

    def createConcepts(self, concept, el, i=10, n=1):
        if i == 0:
            return

        spaces = u" " * n
        i -= 1

        id = el.get(u"id")
        tag = el.tag

        if id is not None:
            c = concept.addConceptKeyType(id, tag)
        else:
            c = concept.addConceptKeyType(tag, tag)

        logger.info(u"%s%s[%s]" % (spaces, c.name, c.typeName))

        self.conceptAttributes(c, el, n+1)

        for elm in el:
            self.createConcepts(c, elm, i, n+1)

    def createArchimate(self, fileArchiModel, fileArchiP):
        archi = Concepts.loadConcepts(fileArchiP)

        rootName = etree.QName(ARCHIMATE_NS, u'model')
        root = etree.Element(rootName, version=u"2.6.0", name=fileArchiP, id=u"02cec69f", nsmap=NS_MAP)
        xmlSheet = etree.ElementTree(root)

        self.createArchimateElements(xmlSheet, archi, root)

        output = StringIO.StringIO()
        xmlSheet.write(output, pretty_print=True)

        logger.info(u"%s" % (output.getvalue()))

        f = open(fileArchiModel, u'w')
        f.write(output.getvalue())
        f.close()

        output.close()

    def createArchimateElements(self, xmlSheet, archi, root, n=1):

        spaces = u" " * n

        cd = archi.getConcepts().values()

        for x in cd:
            logger.debug(u"%s%s:%s" % (spaces, x.typeName, x.name))

            if x.typeName != u"Attribute":

                tag = x.typeName
                id = x.name
                attrib = x.getConcepts()[u"Attributes"]

                ad = dict()
                for y in attrib.getConcepts().values():
                    for z in attrib.getConcepts().values():
                        ad[z.name] = z.typeName

                element = etree.SubElement(root, tag, ad)

                self.createArchimateElements(xmlSheet, x, element)



    def _checkDuplicate(self, dmID, x):
        xp = u"//element[@id='" + dmID + u"']"
        dm = self.tree.xpath(xp)[0]

        dml = dm.getchildren()

        Duplicate = False
        for xdml in dml:
            logger.info(u"%s" % (xdml.get(u"name")))
            xdml_name = xdml.get(u"name")
            if xdml_name == x.name:
                logger.debug(u"%s Duplicate!" % xdml.get(u"id"))
                return xdml.get(u"id")

        logger.debug(u"dml[%d]" % (len(dml)))

        return None

    def _getID(self):
        r = str(hex(random.randint(0, 16777215)))[-6:] + str(hex(random.randint(0, 16777215))[-2:])

        if r[0] == u"x":
            return self.getID()
        return r

    def getID(self):
        return self._getID()

    def _cleanString(self, s):
        r = u""
        if s == None:
            return r

        for x in s.lstrip(u" "):
            if x.isalnum() or x in (u" ", u"-", u"."):
                r = r + x
        return r.lstrip(u" ").rstrip(u" ")

    def cleanString(self, s):
        return  self._cleanString(s)


    def _cleanCapital(self, s):
        r = ""

        if s == None:
            return None

        v = s.replace(u":", u" ")
        v = v.replace(u"|", u" ")
        v = v.replace(u" ", u" ")
        v = v.replace(u"_", u" ")
        v = v.replace(u"-", u" ")
        v = v.replace(u"/", u" ")

        n = 0
        for x in v:
            if x == x.upper() and n != 0:
                r = r + u" " + x
            else:
                r =r + x

            n += 1

        logger.debug(u"cleanCapital : %s" % r)
        return r

    def addToNodeDict(self, type, d):
        if type in d:
            d[type] += 1
        else:
            d[type] = 1

    def cleanConcept(self, concept):
        if concept.typeName[:10] == u"archimate:" :
            concept.typeName = concept.typeName[10:]

        concept.name = concept.name.replace(u"\"", u"'")

        return concept

    @staticmethod
    def startTimer():
        # measure process time, wall time
        t0 = time.clock()
        start_time = time.time()
        strStartTime = time.asctime(time.localtime(start_time))
        logger.info(u"Start time : %s" % strStartTime)

        return start_time

    @staticmethod
    def stopTimer(start_time):
        #measure wall time
        strStartTime = time.asctime(time.localtime(start_time))
        logger.info(u"Start time : %s" % strStartTime)

        end_time = time.time()

        strEndTime = time.asctime(time.localtime(end_time))
        logger.info(u"End   time : %s" % strEndTime)

        # measure process time
        timeTaken = end_time - start_time

        minutes = timeTaken % 60
        hours = minutes / 60
        seconds = timeTaken - (minutes * 60.0)
        logger.info(u"Process Time = %4.2f seconds, %d Minute(s), %d hours" % (timeTaken, minutes, hours))

if __name__ == u"__main__":
    fileArchimate = u"test" + os.sep + u"Testing.archimate"

    al = ArchiLib()

    al.logTypeCounts()