#! /usr/bin/python
#
# Estimate Scope_Items Export
#
__author__ = 'morrj140'
__VERSION__ = '0.1'

import datetime
import time
import logging
from nl_lib import Logger
from nl_lib.ConceptGraph import Neo4JGraph
from nl_lib.Concepts import Concepts
from nl_lib.Constants import *

logger = Logger.setupLogging(__name__)
logger.setLevel(logging.INFO)

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
from openpyxl.worksheet import Worksheet as worksheet

from al_ArchiLib import *
import al_QueryGraph as QG

def queryExportExcel(lq, fileIn, fileOut, wst=None):

    wb = load_workbook(filename = fileIn)

    if wst == None:
        wsTitle = "Import Items" + time.strftime("%Y%d%m_%H%M%S")
    else:
        wsTitle = wst

    ws = wb.create_sheet()
    ws.title = wsTitle

    logger.info("Created Worksheet %s" % wsTitle)

    n = 0
    for x in lq:
        n += 1

        logger.debug("x in lq : %s" % x)

        m = 0
        for y in x:
            m += 1

            col = get_column_letter(m)
            logger.debug("col : %s" % col)

            logger.debug("y : %s" % y)

            rs = "%s%s" % (col, n)
            logger.debug("%s : %s" % (rs, y))

            ws.cell(rs).value = ("%s" % (y))

    wb.save(filename = fileOut)

    logger.info("Saved file : %s" % fileOut)

if __name__ == "__main__":
    fileIn = 'Template_Estimate.xlsx'
    #fileOut = 'Template_Estimate_%s_new.xlsx' % time.strftime("%Y%d%m_%H%M%S")
    fileOut = 'Template_Estimate_new.xlsx'

    graph = Neo4JGraph(gdb)

    qs = "MATCH "
    qs = qs +    "(n0:ApplicationFunction)-- (r0)"
    qs = qs + "-- (n1:ApplicationComponent)--(r1)"
    qs = qs + "-- (n2:ApplicationService)--  (r2)"
    qs = qs + "-- (n3:BusinessProcess)--     (r3)"
    qs = qs + "-- (n4:BusinessObject) "
    qs = qs + "Return n0, r0, n1, r1, n2, r2, n3, r3, n4, n4.PageRank, n4.RequirementCount, n4.Degree"

    lq, qd = QG.cypherQuery(graph, qs)

    queryExportExcel(lq, fileIn, fileOut)

    logger.info("%d rows returned" % len(lq))