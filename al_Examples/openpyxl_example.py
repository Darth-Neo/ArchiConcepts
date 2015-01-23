__author__ = 'morrj140'

import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter

def exportExcel(fileIn, fileOut):

    #wb = Workbook()
    wb = load_workbook(filename = fileIn)

    #ws = wb.active
    ws = wb.create_sheet()

    ws.title = "Scope Items"
    ws.title = 'Pi'
    ws['F5'] = 3.14

    for col_idx in range(1, 40):
        col = get_column_letter(col_idx)

        for row in range(1, 600):
            ws.cell('%s%s'%(col, row)).value = '%s%s' % (col, row)

    wb.save(filename = fileOut)

if __name__ == "__main__":
    fileIn = '/Users/morrj140/Development/GitRepository/ArchiConcepts/export.xlsx'
    fileOut = '/Users/morrj140/Development/GitRepository/ArchiConcepts/export_new.xlsx'

    exportExcel(fileIn, fileOut)